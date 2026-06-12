import json
import logging
import os
import re
import sys
import subprocess
import time
import uuid
from collections.abc import AsyncIterator
from io import BytesIO
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

import httpx
from pypdf import PdfReader
import xlrd
from openpyxl import load_workbook
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from loguru import logger
from pydantic import BaseModel
from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parent
CONTEXT_FILE = BASE_DIR / "岗位标准化母体.history.json"
STATIC_DIR = BASE_DIR / "static"
UPLOAD_ROOT = BASE_DIR / "uploads"
SUPPORTED_ATTACHMENT_EXTS = (".txt", ".md", ".json", ".csv", ".doc", ".docx", ".xls", ".xlsx", ".pdf")

# Load environment variables from local .env before reading os.getenv.
load_dotenv(dotenv_path=BASE_DIR / ".env")


class InterceptHandler(logging.Handler):
    def emit(self, record: logging.LogRecord) -> None:
        try:
            level: str | int = logger.level(record.levelname).name
        except ValueError:
            level = record.levelno

        frame = logging.currentframe()
        depth = 2
        while frame and frame.f_code.co_filename == logging.__file__:
            frame = frame.f_back
            depth += 1

        logger.opt(depth=depth, exception=record.exc_info).log(level, record.getMessage())


def _setup_logging() -> None:
    log_level = os.getenv("LOG_LEVEL", "INFO").upper()
    log_file = os.getenv("LOG_FILE", "app.log").strip() or "app.log"
    log_rotation = os.getenv("LOG_ROTATION", "1 day").strip() or "1 day"
    log_retention = os.getenv("LOG_RETENTION", "14 days").strip() or "14 days"
    log_json = os.getenv("LOG_JSON", "false").strip().lower() in {"1", "true", "yes", "on"}
    log_dir = BASE_DIR / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)

    logger.remove()
    logger.add(
        sink=sys.stderr,
        level=log_level,
        colorize=not log_json,
        serialize=log_json,
        backtrace=True,
        diagnose=False,
    )
    logger.add(
        log_dir / log_file,
        level=log_level,
        rotation=log_rotation,
        retention=log_retention,
        encoding="utf-8",
        enqueue=True,
        serialize=log_json,
        backtrace=True,
        diagnose=False,
    )

    intercept_handler = InterceptHandler()
    logging.basicConfig(handlers=[intercept_handler], level=0, force=True)
    for logger_name in ("uvicorn", "uvicorn.error", "uvicorn.access", "fastapi"):
        logging_logger = logging.getLogger(logger_name)
        logging_logger.handlers = [intercept_handler]
        logging_logger.propagate = False


_setup_logging()


def _now_iso() -> str:
    return datetime.now(UTC).isoformat()


@dataclass
class ChatMessage:
    role: str
    content: str
    created_at: str = field(default_factory=_now_iso)
    is_context: bool = False
    visible_in_history: bool = True
    attachments: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class ChatSession:
    session_id: str
    show_context_in_history: bool
    context_file: str
    created_at: str = field(default_factory=_now_iso)
    messages: list[ChatMessage] = field(default_factory=list)


class CreateSessionRequest(BaseModel):
    show_context_in_history: bool = True


class UpdateSessionSettingsRequest(BaseModel):
    show_context_in_history: bool


class SessionResponse(BaseModel):
    session_id: str
    show_context_in_history: bool
    created_at: str
    history: list[dict[str, Any]]


class SessionSummaryResponse(BaseModel):
    session_id: str
    created_at: str
    updated_at: str
    latest_preview: str


app = FastAPI(title="Gangbiao Chatbot", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def log_request(request: Request, call_next):
    request_id = request.headers.get("x-request-id") or uuid.uuid4().hex[:12]
    request.state.request_id = request_id
    req_logger = LOGGER.bind(
        request_id=request_id,
        method=request.method,
        path=request.url.path,
    )

    started = time.perf_counter()
    req_logger.info("request_started")

    try:
        response = await call_next(request)
    except Exception:
        elapsed_ms = (time.perf_counter() - started) * 1000
        req_logger.bind(elapsed_ms=round(elapsed_ms, 2)).exception("request_failed")
        raise

    elapsed_ms = (time.perf_counter() - started) * 1000
    response.headers["x-request-id"] = request_id
    req_logger.bind(status_code=response.status_code, elapsed_ms=round(elapsed_ms, 2)).info("request_finished")
    return response

if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

SESSIONS: dict[str, ChatSession] = {}
LOGGER = logger.bind(component="chatbot")
MATERIALS_CONTEXT_CACHE: list[ChatMessage] = []


def _env_flag(name: str, default: bool) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "on"}


def _resolve_materials_dir() -> Path | None:
    raw = os.getenv("MATERIALS_DIR", "").strip()
    if not raw:
        return None

    p = Path(raw).expanduser()
    if not p.is_absolute():
        p = (BASE_DIR / p).resolve()
    return p


def _clone_context_messages(messages: list[ChatMessage]) -> list[ChatMessage]:
    cloned: list[ChatMessage] = []
    for msg in messages:
        cloned.append(
            ChatMessage(
                role=msg.role,
                content=msg.content,
                is_context=msg.is_context,
                visible_in_history=msg.visible_in_history,
                attachments=[dict(a) for a in msg.attachments],
            )
        )
    return cloned


def _load_materials_context_messages() -> list[ChatMessage]:
    global MATERIALS_CONTEXT_CACHE

    if not _env_flag("MATERIALS_AUTOLOAD", True):
        return []

    if MATERIALS_CONTEXT_CACHE:
        return _clone_context_messages(MATERIALS_CONTEXT_CACHE)

    materials_dir = _resolve_materials_dir()
    if not materials_dir:
        return []

    if not materials_dir.exists() or not materials_dir.is_dir():
        LOGGER.warning("MATERIALS_DIR does not exist or is not a directory: {}", materials_dir)
        return []

    max_files = max(int(os.getenv("MATERIALS_MAX_FILES", "20") or "20"), 1)
    max_excerpt_chars = max(int(os.getenv("MATERIALS_MAX_EXCERPT_CHARS", "1200") or "1200"), 200)

    candidates = sorted(
        [
            p for p in materials_dir.rglob("*")
            if p.is_file() and p.suffix.lower() in SUPPORTED_ATTACHMENT_EXTS
        ],
        key=lambda p: str(p),
    )[:max_files]

    if not candidates:
        LOGGER.info("No supported materials found in MATERIALS_DIR: {}", materials_dir)
        return []

    blocks: list[str] = []
    attachments: list[dict[str, Any]] = []

    for path in candidates:
        try:
            raw = path.read_bytes()
        except OSError as exc:
            LOGGER.warning("Failed to read material file {}: {}", path, exc)
            continue

        excerpt = _extract_attachment_excerpt(raw_bytes=raw, lower_name=path.name.lower())
        if not excerpt:
            continue

        rel = str(path.relative_to(BASE_DIR)) if path.is_relative_to(BASE_DIR) else str(path)
        attachments.append(
            {
                "filename": path.name,
                "size": len(raw),
                "saved_path": rel,
                "original_path": str(path),
                "source": "materials_dir",
            }
        )
        blocks.append(
            f"[教材] {path.name}\n"
            f"路径: {path}\n"
            f"摘要:\n{excerpt[:max_excerpt_chars]}"
        )

    if not blocks:
        LOGGER.info("No readable excerpts extracted from MATERIALS_DIR: {}", materials_dir)
        return []

    content = (
        f"以下内容来自自动加载教材目录: {materials_dir}\n"
        f"已加载 {len(blocks)} 份材料（最多 {max_files} 份）。\n\n"
        + "\n\n---\n\n".join(blocks)
    )

    MATERIALS_CONTEXT_CACHE = [
        ChatMessage(
            role="system",
            content=content,
            is_context=True,
            visible_in_history=False,
            attachments=attachments,
        )
    ]

    LOGGER.info("Loaded {} materials into default context from {}", len(blocks), materials_dir)
    return _clone_context_messages(MATERIALS_CONTEXT_CACHE)


def _load_default_context_messages() -> list[ChatMessage]:
    if not CONTEXT_FILE.exists():
        LOGGER.warning("Default context file not found: {}", CONTEXT_FILE)
        return []

    raw = json.loads(CONTEXT_FILE.read_text(encoding="utf-8"))
    messages: list[ChatMessage] = []

    header = {
        "version": raw.get("version"),
        "format": raw.get("format"),
        "generated_at": raw.get("generated_at"),
        "source_file": CONTEXT_FILE.name,
    }
    LOGGER.info("Default context metadata loaded: {}", json.dumps(header, ensure_ascii=False))

    for item in raw.get("messages", []):
        role = item.get("role", "user")
        content = item.get("content", "")
        meta = item.get("metadata") if isinstance(item.get("metadata"), dict) else None
        attachments = _attachments_from_history_metadata(meta)
        payload = content
        summary = _compact_metadata_summary(meta)
        if summary:
            payload = (
                f"{content}\n\n[metadata]\n"
                f"{json.dumps(summary, ensure_ascii=False)}"
            )
        messages.append(
            ChatMessage(
                role=role,
                content=payload,
                is_context=True,
                attachments=attachments,
            )
        )

    return messages


def _attachments_from_history_metadata(meta: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not meta:
        return []

    raw = meta.get("attachments")
    if not isinstance(raw, list):
        return []

    normalized: list[dict[str, Any]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue

        filename = str(item.get("filename") or "unnamed")
        original_path = item.get("original_path")
        normalized.append(
            {
                "filename": filename,
                "content_type": item.get("content_type"),
                "size": item.get("size"),
                "saved_path": item.get("saved_path"),
                "original_path": str(original_path) if original_path else None,
                "source": "history_context",
            }
        )

    return normalized


def _compact_metadata_summary(meta: dict[str, Any] | None) -> dict[str, Any]:
    if not meta:
        return {}

    summary: dict[str, Any] = {}

    if "attachments" in meta:
        attachments = _attachments_from_history_metadata(meta)
        if attachments:
            summary["attachments"] = [
                {
                    "filename": item.get("filename"),
                    "original_path": item.get("original_path"),
                }
                for item in attachments
            ]

    if "details" in meta and isinstance(meta.get("details"), list):
        details = [
            str(d.get("summary"))
            for d in meta["details"]
            if isinstance(d, dict) and d.get("summary")
        ]
        if details:
            summary["details_summary"] = details[:6]

    for key in ("source", "tool", "stage", "note"):
        if key in meta and meta[key] not in (None, ""):
            summary[key] = meta[key]

    return summary


def _session_history_for_client(session: ChatSession) -> list[dict[str, Any]]:
    history: list[dict[str, Any]] = []
    for msg in session.messages:
        if not msg.visible_in_history:
            continue
        if msg.is_context and not session.show_context_in_history:
            continue
        history.append(
            {
                "role": msg.role,
                "content": msg.content,
                "created_at": msg.created_at,
                "is_context": msg.is_context,
                "attachments": msg.attachments,
            }
        )
    return history


def _session_summary_for_client(session: ChatSession) -> dict[str, str]:
    last_message = session.messages[-1] if session.messages else None
    latest_preview = "新会话"

    if last_message:
        preview = last_message.content.strip().replace("\n", " ")
        latest_preview = preview[:40] if preview else "空消息"

    return {
        "session_id": session.session_id,
        "created_at": session.created_at,
        "updated_at": last_message.created_at if last_message else session.created_at,
        "latest_preview": latest_preview,
    }


def _build_model_messages(session: ChatSession, user_msg: ChatMessage) -> list[dict[str, str]]:
    messages: list[dict[str, str]] = [
        {
            "role": "system",
            "content": (
                "你是岗位标准化 AI 教练。"
                "请在回答中保持教练式引导，优先围绕用户提供的上下文和材料。"
            ),
        }
    ]

    for msg in session.messages:
        if msg.role not in {"system", "user", "assistant"}:
            continue
        messages.append({"role": msg.role, "content": msg.content})

    messages.append({"role": "user", "content": user_msg.content})
    return messages


async def _call_llm(messages: list[dict[str, str]]) -> str:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
    base_url = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1").rstrip("/")

    if not api_key:
        last_user = next((m["content"] for m in reversed(messages) if m["role"] == "user"), "")
        return (
            "当前未配置 OPENAI_API_KEY，已使用本地回退回复。\n"
            "如需真实模型回复，请设置环境变量后重试。\n\n"
            f"你刚才的问题是：{last_user[:400]}"
        )

    url = f"{base_url}/chat/completions"
    payload = {"model": model, "messages": messages, "temperature": 0.2}
    headers = {"Authorization": f"Bearer {api_key}"}

    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.post(url, json=payload, headers=headers)
        if response.status_code >= 400:
            raise HTTPException(
                status_code=502,
                detail=f"LLM 调用失败: {response.status_code} {response.text}",
            )
        data = response.json()
        return data["choices"][0]["message"]["content"]


async def _call_llm_stream(messages: list[dict[str, str]]) -> AsyncIterator[str]:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
    base_url = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1").rstrip("/")

    if not api_key:
        last_user = next((m["content"] for m in reversed(messages) if m["role"] == "user"), "")
        fallback = (
            "当前未配置 OPENAI_API_KEY，已使用本地回退回复。\n"
            "如需真实模型回复，请设置环境变量后重试。\n\n"
            f"你刚才的问题是：{last_user[:400]}"
        )
        for i in range(0, len(fallback), 30):
            yield fallback[i:i + 30]
        return

    url = f"{base_url}/chat/completions"
    payload = {"model": model, "messages": messages, "temperature": 0.2, "stream": True}
    headers = {"Authorization": f"Bearer {api_key}", "Accept": "text/event-stream"}

    async with httpx.AsyncClient(timeout=120.0) as client:
        async with client.stream("POST", url, json=payload, headers=headers) as response:
            if response.status_code >= 400:
                body = await response.aread()
                raise HTTPException(
                    status_code=502,
                    detail=f"LLM 调用失败: {response.status_code} {body.decode('utf-8', errors='ignore')}",
                )

            async for raw_line in response.aiter_lines():
                line = raw_line.strip()
                if not line.startswith("data:"):
                    continue

                data_line = line[5:].strip()
                if data_line == "[DONE]":
                    break

                try:
                    data = json.loads(data_line)
                except json.JSONDecodeError:
                    continue

                choice = (data.get("choices") or [{}])[0]
                delta = (choice.get("delta") or {}).get("content")

                if isinstance(delta, str) and delta:
                    yield delta
                    continue

                # Compatibility fallback for gateways that emit full message chunks.
                content = ((choice.get("message") or {}).get("content"))
                if isinstance(content, str) and content:
                    yield content


async def _save_attachments(
    session_id: str,
    files: list[UploadFile],
) -> tuple[list[dict[str, Any]], list[str]]:
    saved_meta: list[dict[str, Any]] = []
    attachment_hints: list[str] = []

    if not files:
        return saved_meta, attachment_hints

    UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    session_dir = UPLOAD_ROOT / session_id
    session_dir.mkdir(parents=True, exist_ok=True)

    for f in files:
        raw_bytes = await f.read()
        filename = f.filename or "unnamed"
        safe_name = Path(filename).name
        target = session_dir / f"{uuid.uuid4().hex}_{safe_name}"
        target.write_bytes(raw_bytes)

        lower_name = safe_name.lower()
        excerpt = _extract_attachment_excerpt(raw_bytes=raw_bytes, lower_name=lower_name)

        meta = {
            "filename": safe_name,
            "content_type": f.content_type,
            "size": len(raw_bytes),
            "saved_path": str(target.relative_to(BASE_DIR)),
            "excerpt": excerpt[:800],
        }
        saved_meta.append(meta)

        hint = f"附件: {safe_name} ({len(raw_bytes)} bytes)"
        if excerpt:
            hint += f"\n可读摘要:\n{excerpt[:500]}"
        else:
            hint += "\n未提取到可读文本（建议使用 txt/md/json/csv/doc/docx/xls/xlsx/pdf，或粘贴关键内容）"
        attachment_hints.append(hint)

    return saved_meta, attachment_hints


def _extract_attachment_excerpt(raw_bytes: bytes, lower_name: str) -> str:
    if lower_name.endswith((".txt", ".md", ".json", ".csv")):
        return raw_bytes[:4000].decode("utf-8", errors="ignore")

    if lower_name.endswith(".docx"):
        return _extract_docx_text(raw_bytes)[:4000]

    if lower_name.endswith(".doc"):
        return _extract_doc_text(raw_bytes)[:4000]

    if lower_name.endswith(".pdf"):
        return _extract_pdf_text(raw_bytes)[:4000]

    if lower_name.endswith(".xlsx"):
        return _extract_xlsx_text(raw_bytes)[:4000]

    if lower_name.endswith(".xls"):
        return _extract_xls_text(raw_bytes)[:4000]

    return ""


def _extract_docx_text(raw_bytes: bytes) -> str:
    try:
        with ZipFile(BytesIO(raw_bytes)) as zf:
            xml_bytes = zf.read("word/document.xml")
        root = ElementTree.fromstring(xml_bytes)
    except (BadZipFile, KeyError, ElementTree.ParseError, ValueError):
        return ""

    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []

    for paragraph in root.findall(".//w:p", ns):
        text = "".join(paragraph.itertext()).strip()
        if text:
            paragraphs.append(text)

    if not paragraphs:
        text = "".join(root.itertext()).strip()
        return text

    return "\n".join(paragraphs)


def _extract_pdf_text(raw_bytes: bytes) -> str:
    try:
        reader = PdfReader(BytesIO(raw_bytes))
    except Exception:
        return ""

    chunks: list[str] = []
    for page in reader.pages:
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        text = text.strip()
        if text:
            chunks.append(text)

    return "\n".join(chunks)


def _extract_doc_text(raw_bytes: bytes) -> str:
    # Prefer antiword when available, then fall back to binary string extraction.
    try:
        with NamedTemporaryFile(suffix=".doc", delete=True) as tmp:
            tmp.write(raw_bytes)
            tmp.flush()
            completed = subprocess.run(
                ["antiword", "-w", "0", tmp.name],
                capture_output=True,
                text=True,
                timeout=10,
                check=False,
            )
            if completed.returncode == 0 and completed.stdout.strip():
                return completed.stdout
    except (FileNotFoundError, subprocess.SubprocessError, OSError):
        pass

    return _extract_printable_strings(raw_bytes)


def _extract_xlsx_text(raw_bytes: bytes) -> str:
    try:
        # Some files have incorrect dimension metadata (e.g. A1), which breaks
        # read_only iteration. Normal mode is more robust for these workbooks.
        wb = load_workbook(filename=BytesIO(raw_bytes), data_only=False, read_only=False)
    except Exception:
        return ""

    lines: list[str] = []
    max_rows_limit = 1000
    max_cols_limit = 80

    for sheet in wb.worksheets:
        lines.append(f"[Sheet] {sheet.title}")
        merged_values: dict[tuple[int, int], Any] = {}
        for merged in sheet.merged_cells.ranges:
            top_left = sheet.cell(row=merged.min_row, column=merged.min_col).value
            if top_left in (None, ""):
                continue
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    merged_values[(r, c)] = top_left

        max_rows = min(max(sheet.max_row, 1), max_rows_limit)
        max_cols = min(max(sheet.max_column, 1), max_cols_limit)

        for row_idx in range(1, max_rows + 1):
            cells: list[str] = []
            row_has_value = False

            for col_idx in range(1, max_cols + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if value in (None, ""):
                    value = merged_values.get((row_idx, col_idx))

                normalized = _normalize_cell_value(value) if value not in (None, "") else ""
                if normalized:
                    row_has_value = True
                cells.append(normalized)

            if row_has_value:
                # Trim trailing empty columns but keep interior empties for table shape.
                while cells and cells[-1] == "":
                    cells.pop()
                lines.append("\t".join(cells))

        if sheet.max_row > max_rows or sheet.max_column > max_cols:
            lines.append("...")

    return _normalize_gangbiao_labels("\n".join(lines))


def _extract_xls_text(raw_bytes: bytes) -> str:
    try:
        book = xlrd.open_workbook(file_contents=raw_bytes)
    except Exception:
        return ""

    lines: list[str] = []
    max_rows_limit = 1000
    max_cols_limit = 80

    for sheet in book.sheets():
        lines.append(f"[Sheet] {sheet.name}")
        max_rows = min(max(sheet.nrows, 1), max_rows_limit)
        max_cols = min(max(sheet.ncols, 1), max_cols_limit)

        merged_values: dict[tuple[int, int], Any] = {}
        for rlo, rhi, clo, chi in getattr(sheet, "merged_cells", []):
            top_left = sheet.cell_value(rlo, clo)
            if top_left in (None, ""):
                continue
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    merged_values[(r, c)] = top_left

        for r in range(max_rows):
            cells: list[str] = []
            row_has_value = False
            for c in range(max_cols):
                value = sheet.cell_value(r, c)
                if value in (None, ""):
                    value = merged_values.get((r, c))

                normalized = _normalize_cell_value(value) if value not in (None, "") else ""
                if normalized:
                    row_has_value = True
                cells.append(normalized)

            if row_has_value:
                while cells and cells[-1] == "":
                    cells.pop()
                lines.append("\t".join(cells))

        if sheet.nrows > max_rows or sheet.ncols > max_cols:
            lines.append("...")

    return _normalize_gangbiao_labels("\n".join(lines))


def _normalize_gangbiao_labels(text: str) -> str:
    if not text:
        return text

    normalized_lines: list[str] = []
    for raw_line in text.splitlines():
        line = raw_line

        # Normalize common purpose label variants.
        line = re.sub(r"\b目的\s*[:：]", "任务目的：", line)

        # Normalize result label variants while keeping the remaining sentence.
        line = re.sub(
            r"\b成果\s*(?:（[^）]*）)?\s*[:：]",
            "任务成果（预算、交期、完成度）：",
            line,
        )

        normalized_lines.append(line)

    return "\n".join(normalized_lines)


def _extract_printable_strings(raw_bytes: bytes) -> str:
    text_parts: list[str] = []

    ascii_runs = re.findall(rb"[\x20-\x7e\t\r\n]{6,}", raw_bytes)
    if ascii_runs:
        text_parts.append("\n".join(run.decode("latin-1", errors="ignore") for run in ascii_runs[:200]))

    utf16 = raw_bytes.decode("utf-16-le", errors="ignore")
    utf16_clean = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", utf16)
    utf16_lines = [line.strip() for line in utf16_clean.splitlines() if len(line.strip()) >= 6]
    if utf16_lines:
        text_parts.append("\n".join(utf16_lines[:200]))

    return "\n".join(part for part in text_parts if part).strip()


def _normalize_cell_value(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@app.get("/")
async def index() -> FileResponse:
    html = STATIC_DIR / "index.html"
    if not html.exists():
        raise HTTPException(status_code=404, detail="前端页面不存在")
    return FileResponse(html)


@app.get("/api/health")
async def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/api/sessions", response_model=SessionResponse)
async def create_session(req: CreateSessionRequest) -> SessionResponse:
    sid = uuid.uuid4().hex
    session = ChatSession(
        session_id=sid,
        show_context_in_history=req.show_context_in_history,
        context_file=CONTEXT_FILE.name,
    )
    session.messages.extend(_load_default_context_messages())
    session.messages.extend(_load_materials_context_messages())
    SESSIONS[sid] = session
    LOGGER.bind(session_id=sid).info(
        "session_created show_context_in_history={} message_count={}",
        session.show_context_in_history,
        len(session.messages),
    )
    return SessionResponse(
        session_id=sid,
        show_context_in_history=session.show_context_in_history,
        created_at=session.created_at,
        history=_session_history_for_client(session),
    )


@app.get("/api/sessions", response_model=list[SessionSummaryResponse])
async def list_sessions() -> list[dict[str, str]]:
    summaries = [_session_summary_for_client(session) for session in SESSIONS.values()]
    return sorted(summaries, key=lambda item: item["updated_at"], reverse=True)


@app.patch("/api/sessions/{session_id}/settings", response_model=SessionResponse)
async def update_session_settings(
    session_id: str,
    req: UpdateSessionSettingsRequest,
) -> SessionResponse:
    session = SESSIONS.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")
    session.show_context_in_history = req.show_context_in_history
    return SessionResponse(
        session_id=session.session_id,
        show_context_in_history=session.show_context_in_history,
        created_at=session.created_at,
        history=_session_history_for_client(session),
    )


@app.get("/api/sessions/{session_id}", response_model=SessionResponse)
async def get_session(session_id: str) -> SessionResponse:
    session = SESSIONS.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")
    return SessionResponse(
        session_id=session.session_id,
        show_context_in_history=session.show_context_in_history,
        created_at=session.created_at,
        history=_session_history_for_client(session),
    )


@app.post("/api/chat")
async def chat(
    session_id: str = Form(...),
    message: str = Form(...),
    files: list[UploadFile] = File(default=[]),
) -> dict[str, Any]:
    session = SESSIONS.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    user_text = message.strip()
    if not user_text and not files:
        raise HTTPException(status_code=400, detail="消息和附件不能同时为空")

    saved_files, file_hints = await _save_attachments(session_id=session_id, files=files)
    chat_logger = LOGGER.bind(session_id=session_id)
    chat_logger.info(
        "chat_request_received attachments={} has_text={}",
        len(saved_files),
        bool(user_text),
    )
    final_user_text = user_text
    if file_hints:
        final_user_text = f"{user_text}\n\n" + "\n\n".join(file_hints)

    user_msg = ChatMessage(role="user", content=final_user_text, attachments=saved_files)
    session.messages.append(user_msg)

    llm_messages = _build_model_messages(session, user_msg)
    assistant_text = await _call_llm(llm_messages)

    assistant_msg = ChatMessage(role="assistant", content=assistant_text)
    session.messages.append(assistant_msg)
    chat_logger.info("chat_reply_generated reply_chars={}", len(assistant_text))

    return {
        "session_id": session.session_id,
        "reply": assistant_text,
        "history": _session_history_for_client(session),
    }


@app.post("/api/chat/stream")
async def chat_stream(
    session_id: str = Form(...),
    message: str = Form(...),
    files: list[UploadFile] = File(default=[]),
) -> StreamingResponse:
    session = SESSIONS.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="会话不存在")

    user_text = message.strip()
    if not user_text and not files:
        raise HTTPException(status_code=400, detail="消息和附件不能同时为空")

    saved_files, file_hints = await _save_attachments(session_id=session_id, files=files)
    stream_logger = LOGGER.bind(session_id=session_id)
    stream_logger.info(
        "chat_stream_request_received attachments={} has_text={}",
        len(saved_files),
        bool(user_text),
    )
    final_user_text = user_text
    if file_hints:
        final_user_text = f"{user_text}\n\n" + "\n\n".join(file_hints)

    user_msg = ChatMessage(role="user", content=final_user_text, attachments=saved_files)
    session.messages.append(user_msg)
    llm_messages = _build_model_messages(session, user_msg)

    async def event_gen() -> AsyncIterator[str]:
        chunks: list[str] = []
        try:
            async for delta in _call_llm_stream(llm_messages):
                chunks.append(delta)
                yield f"data: {json.dumps({'type': 'delta', 'delta': delta}, ensure_ascii=False)}\n\n"
        except HTTPException as exc:
            stream_logger.warning("chat_stream_http_error detail={}", exc.detail)
            yield f"data: {json.dumps({'type': 'error', 'error': str(exc.detail)}, ensure_ascii=False)}\n\n"
            return
        except Exception as exc:  # pragma: no cover - defensive branch
            stream_logger.exception("chat_stream_unexpected_error")
            yield f"data: {json.dumps({'type': 'error', 'error': f'流式输出失败: {exc}'}, ensure_ascii=False)}\n\n"
            return

        assistant_text = "".join(chunks).strip() or "（模型未返回内容）"
        assistant_msg = ChatMessage(role="assistant", content=assistant_text)
        session.messages.append(assistant_msg)
        stream_logger.info("chat_stream_reply_generated reply_chars={} chunk_count={}", len(assistant_text), len(chunks))

        done_payload = {
            "type": "done",
            "reply": assistant_text,
            "history": _session_history_for_client(session),
        }
        yield f"data: {json.dumps(done_payload, ensure_ascii=False)}\n\n"

    return StreamingResponse(event_gen(), media_type="text/event-stream")


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=2088, reload=True)