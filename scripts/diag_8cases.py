#!/usr/bin/env python3
"""List the multi-sheet "selection problem" failures for manual review.

A case is a "selection problem" when the sheet the extractor currently picks
has a stage fingerprint that does NOT match the expected label's stage
(i.e. the wrong sheet is being read). Contrast with "version confusion"
where the right sheet is read but D2/D6 or D3/D7 is judged wrong.
"""
import json
from pathlib import Path
from io import BytesIO
from openpyxl import load_workbook

CASES = Path("tests/fixtures/classifier/cases")
LABELS = json.loads(Path("tests/fixtures/classifier/labels.json").read_text(encoding="utf-8"))

# (stage_id, stage_name, [fingerprint substrings])
STAGE_FPS = [
    ("一", "阶段一【岗位价值和岗位任务】", ["岗位价值（对客户而言）", "岗位任务的目的", "【岗位价值和岗位任务】教材链接"]),
    ("二", "阶段二【工时管理】", ["核心任务实际时间投入", "执行工时统计", "【工时管理】教材链接"]),
    ("三", "阶段三【人效管理】", ["建立人效模型", "人效模型", "【人效管理】教材链接"]),
    ("四", "阶段四【关键活动工作分解表】", ["关键活动名称", "列为关键活动的理由", "关键活动的目的和成果", "【工作分解表】教材链接"]),
]
LABEL_STAGE = {"D1": "一", "D5": "一", "D2": "二", "D6": "二", "D3": "三", "D7": "三", "D4": "四"}
STAGE_NAME = {"一": "阶段一", "二": "阶段二", "三": "阶段三", "四": "阶段四"}

def sheet_stage(blob: str) -> str | None:
    for sid, _, fps in STAGE_FPS:
        if any(fp in blob for fp in fps):
            return sid
    return None

def pick_sheet(raw: bytes):
    """Replicate _collect_xlsx sheet selection."""
    wb = load_workbook(filename=BytesIO(raw), data_only=False, read_only=False)
    sheet = None
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        if any(cell.value not in (None, "") for row in ws.iter_rows() for cell in row):
            sheet = ws
            break
    if sheet is None:
        for ws in wb.worksheets:
            if any(cell.value not in (None, "") for row in ws.iter_rows() for cell in row):
                sheet = ws
                break
    if sheet is None:
        sheet = wb.worksheets[0]
    return wb, sheet

selection_problems = []
version_confusions = []

for fname, expected in LABELS.items():
    p = CASES / fname
    if not p.exists() or not fname.endswith(".xlsx"):
        continue
    raw = p.read_bytes()
    wb, picked = pick_sheet(raw)
    picked_blob = "\n".join(str(v) for row in picked.iter_rows(values_only=True) for v in row if isinstance(v, str) and v.strip())
    picked_stage = sheet_stage(picked_blob)
    exp_stage = LABEL_STAGE.get(expected.upper())

    # build per-sheet info
    sheets_info = []
    for i, ws in enumerate(wb.worksheets):
        blob = "\n".join(str(v) for row in ws.iter_rows(values_only=True) for v in row if isinstance(v, str) and v.strip())
        st = sheet_stage(blob)
        nonempty = "有内容" if blob else "空表"
        vis = "可见" if ws.sheet_state == "visible" else "隐藏"
        marker = " ←当前选中" if ws.title == picked.title else ""
        target_marker = " ★目标阶段" if st == exp_stage else ""
        sheets_info.append((i + 1, ws.title, vis, f"{ws.max_row}×{ws.max_column}", nonempty, STAGE_NAME.get(st, "无指纹"), marker + target_marker))

    is_sel_problem = (picked_stage != exp_stage)
    rec = {
        "file": fname,
        "expected": expected,
        "exp_stage": STAGE_NAME.get(exp_stage),
        "picked_sheet": picked.title,
        "picked_stage": STAGE_NAME.get(picked_stage, "无指纹"),
        "sheets": sheets_info,
    }
    if is_sel_problem:
        selection_problems.append(rec)
    else:
        version_confusions.append(rec)

print(f"=== 选中错误 sheet（多 sheet 选择问题）：{len(selection_problems)} 例 ===\n")
for i, r in enumerate(selection_problems, 1):
    print(f"【{i}】{r['file']}")
    print(f"    标签={r['expected']}（应={r['exp_stage']}）  当前选中「{r['picked_sheet']}」（其指纹={r['picked_stage']}）")
    print(f"    各 sheet：")
    for n, name, vis, dim, ne, st, mk in r["sheets"]:
        print(f"      #{n}「{name}」[{vis}] {dim} {ne} 指纹={st}{mk}")
    print()

print(f"=== 选对 sheet 但版本判反：{len(version_confusions)} 例 ===")
for r in version_confusions:
    print(f"  {r['file']}  标签={r['expected']}  选中「{r['picked_sheet']}」(指纹={r['picked_stage']} ✓)")
