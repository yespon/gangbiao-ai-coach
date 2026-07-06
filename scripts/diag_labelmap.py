#!/usr/bin/env python3
"""Full label map for manual review: file -> label, expected stage, and the
stage fingerprint of the sheet the (fixed) extractor currently picks.
Lets a reviewer see at a glance whether each label matches the file's content.
"""
import json
from pathlib import Path
from io import BytesIO
from openpyxl import load_workbook

CASES = Path("tests/fixtures/classifier/cases")
LABELS = json.loads(Path("tests/fixtures/classifier/labels.json").read_text(encoding="utf-8"))

LABEL_STAGE = {
    "D1": "阶段一(多等级)", "D5": "阶段一(普通)",
    "D2": "阶段二(多等级)", "D6": "阶段二(普通)",
    "D3": "阶段三(多等级)", "D7": "阶段三(普通)",
    "D4": "阶段四(工作分解表)", "NONE": "非模板",
}
STAGE_OF = {"D1": "一", "D5": "一", "D2": "二", "D6": "二", "D3": "三", "D7": "三", "D4": "四"}

FPS = [
    ("一", ["岗位价值（对客户而言）", "岗位任务的目的", "【岗位价值和岗位任务】教材链接"]),
    ("二", ["核心任务实际时间投入", "执行工时统计", "【工时管理】教材链接"]),
    ("三", ["建立人效模型", "人效模型", "【人效管理】教材链接"]),
    ("四", ["关键活动名称", "列为关键活动的理由", "关键活动的目的和成果", "【工作分解表】教材链接"]),
]

def stage_of(blob: str) -> str:
    for sid, fps in FPS:
        if any(fp in blob for fp in fps):
            return sid
    return "?"

def pick_sheet(raw: bytes):
    wb = load_workbook(filename=BytesIO(raw), data_only=False, read_only=False)
    sheet = None
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        if any(cell.value not in (None, "") for row in ws.iter_rows() for cell in row):
            sheet = ws
            break
    if sheet is None:
        for ws in wb.worksheets:
            if any(cell.value not in (None, "") for row in ws.iter_rows() for cell in row):
                sheet = ws
                break
    if sheet is None:
        sheet = wb.worksheets[0]
    return sheet

print(f"{'#':<3}{'文件':<52}{'标签':<6}{'应属阶段':<16}{'选中sheet阶段':<12}{'是否吻合'}")
print("-" * 100)
rows = list(LABELS.items())
for i, (fname, label) in enumerate(rows, 1):
    p = CASES / fname
    if not p.exists() or not fname.endswith(".xlsx"):
        print(f"{i:<3}{fname[:50]:<52}{label:<6}{LABEL_STAGE.get(label,'?'):<16}{'(无文件)':<12}")
        continue
    sheet = pick_sheet(p.read_bytes())
    blob = "\n".join(str(v) for row in sheet.iter_rows(values_only=True) for v in row if isinstance(v, str) and v.strip())
    picked_stage = stage_of(blob)
    exp_stage = STAGE_OF.get(label.upper(), "?")
    ok = "✓" if picked_stage == exp_stage else "✗ 不符"
    # also list which visible sheets carry which stage, so reviewer sees alternatives
    print(f"{i:<3}{fname[:50]:<52}{label:<6}{LABEL_STAGE.get(label,'?'):<16}{picked_stage:<12}{ok}")
