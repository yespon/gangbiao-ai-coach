import pytest

from app.services import template_classifier as svc
from app.services.template_classifier import ClassificationResult, parse_classification


def test_parse_clean_json():
    raw = (
        '{"matched": true, "document_id": "D1", "version": "多等级版", '
        '"stage": "阶段一【岗位价值和岗位任务】", "confidence": 0.97, '
        '"matched_signals": ["列头含「岗位任职资格等级」"], "reason": "命中D1"}'
    )
    r = parse_classification(raw)
    assert r.matched is True
    assert r.document_id == "D1"
    assert r.version == "多等级版"
    assert r.confidence == 0.97
    assert r.matched_signals == ["列头含「岗位任职资格等级」"]
    assert r.error is None


def test_parse_json_inside_code_fence_with_prose():
    raw = (
        "好的，分类结果如下：\n"
        "```json\n"
        '{"matched": true, "document_id": "D5", "confidence": 0.9, '
        '"matched_signals": [], "reason": "x"}\n'
        "```\n"
        "以上。"
    )
    r = parse_classification(raw)
    assert r.document_id == "D5"
    assert r.matched is True


def test_parse_json_with_trailing_prose_no_fence():
    raw = (
        '{"matched": false, "document_id": null, "confidence": 0.0, '
        '"matched_signals": [], "reason": "无匹配"} 这是一段说明文字'
    )
    r = parse_classification(raw)
    assert r.matched is False
    assert r.document_id is None
    assert r.error is None


def test_parse_malformed_returns_error_result():
    r = parse_classification("这不是JSON")
    assert r.matched is False
    assert r.document_id is None
    assert r.error is not None
    assert "解析失败" in r.reason


def test_coerce_invalid_document_id_becomes_none_and_unmatched():
    r = parse_classification('{"matched": true, "document_id": "D9", "confidence": 1.5}')
    assert r.document_id is None
    assert r.matched is False  # invalid id forces unmatched
    assert r.confidence == 1.0  # clamped to [0, 1]


def test_coerce_lowercase_document_id_uppercased():
    r = parse_classification('{"matched": true, "document_id": "d1", "confidence": -0.2}')
    assert r.document_id == "D1"
    assert r.confidence == 0.0  # clamped


def test_classification_result_defaults():
    r = ClassificationResult()
    assert r.matched is False
    assert r.document_id is None
    assert r.confidence == 0.0
    assert r.matched_signals == []
    assert r.error is None


from io import BytesIO

from openpyxl import Workbook

from app.services.template_classifier import extract_cells_for_classification


def _save_xlsx(worksheet_setup):
    wb = Workbook()
    ws = wb.active
    ws.title = "T"
    worksheet_setup(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extract_formats_header_and_non_empty_cells():
    def setup(ws):
        ws["A1"] = "岗位名称"
        ws["B1"] = "服务客户"

    text = extract_cells_for_classification(_save_xlsx(setup), ".xlsx")
    assert text.startswith("文件尺寸：1行 × 2列")
    assert "单元格内容（[行,列] 内容）：" in text
    assert "[1,1] 岗位名称" in text
    assert "[1,2] 服务客户" in text


def test_extract_preserves_merged_top_left_only_and_internal_spaces():
    def setup(ws):
        ws["A1"] = "岗位名称"
        ws["C1"] = 5.0  # float that is an integer -> "5"
        ws.merge_cells("A2:A3")
        ws["A2"] = "核心任务"  # merged anchor; A3 must NOT appear
        ws["B2"] = "拟认证等级              实际时间投入占比"  # internal spaces preserved
        ws.merge_cells("B2:B3")

    text = extract_cells_for_classification(_save_xlsx(setup), ".xlsx")
    # 3 rows (A1/B2/A3 merged span), 3 cols (A,B,C)
    assert text.startswith("文件尺寸：3行 × 3列")
    assert "[1,1] 岗位名称" in text
    assert "[1,3] 5" in text  # float-int normalized to "5"
    assert "[2,1] 核心任务" in text
    assert "[2,2] 拟认证等级              实际时间投入占比" in text  # internal spaces kept
    assert "[3,1]" not in text  # merged non-anchor skipped
    assert "[3,2]" not in text  # merged non-anchor skipped


def test_extract_normalizes_float_integer_values():
    def setup(ws):
        ws["A1"] = 7.0
        ws["A2"] = 3.5  # genuine float stays as-is
        ws["A3"] = "文本"

    text = extract_cells_for_classification(_save_xlsx(setup), ".xlsx")
    assert "[1,1] 7" in text
    assert "[1,1] 7.0" not in text
    assert "[2,1] 3.5" in text
    assert "[3,1] 文本" in text


def test_extract_uses_first_non_empty_worksheet():
    wb = Workbook()
    wb.active.title = "空表"  # default empty sheet
    ws2 = wb.create_sheet("数据")
    ws2["A1"] = "岗位名称"
    buf = BytesIO()
    wb.save(buf)
    text = extract_cells_for_classification(buf.getvalue(), ".xlsx")
    assert "[1,1] 岗位名称" in text


@pytest.mark.asyncio
async def test_classify_text_sends_prompt_and_parses_result(monkeypatch):
    captured = {}

    async def fake_call(messages):
        captured["messages"] = messages
        return (
            '{"matched": true, "document_id": "D1", "version": "多等级版", '
            '"stage": "阶段一【岗位价值和岗位任务】", "confidence": 0.97, '
            '"matched_signals": ["s1"], "reason": "ok"}'
        )

    monkeypatch.setattr(svc, "_call_llm_json", fake_call)

    result = await svc.classify_text(
        "文件尺寸：1行 × 1列\n单元格内容（[行,列] 内容）：\n[1,1] 岗位名称"
    )

    assert captured["messages"][0]["role"] == "system"
    assert captured["messages"][0]["content"] == svc.CLASSIFIER_PROMPT
    assert captured["messages"][1]["role"] == "user"
    assert "岗位名称" in captured["messages"][1]["content"]
    assert result.document_id == "D1"
    assert result.matched is True


@pytest.mark.asyncio
async def test_classify_file_extracts_then_classifies(monkeypatch):
    captured = {}

    async def fake_call(messages):
        captured["user"] = messages[-1]["content"]
        return (
            '{"matched": true, "document_id": "D4", "version": "通用版", '
            '"stage": "阶段四【关键活动工作分解表】", "confidence": 0.95, '
            '"matched_signals": [], "reason": "ok"}'
        )

    monkeypatch.setattr(svc, "_call_llm_json", fake_call)

    def setup(ws):
        ws["A1"] = "关键活动名称"
        ws["B1"] = "列为关键活动的理由"

    result = await svc.classify_file(_save_xlsx(setup), ".xlsx")

    assert "文件尺寸" in captured["user"]
    assert "关键活动名称" in captured["user"]
    assert result.document_id == "D4"


@pytest.mark.asyncio
async def test_call_llm_json_raises_without_api_key(monkeypatch):
    monkeypatch.setattr(svc.settings, "openai_api_key", "")
    with pytest.raises(RuntimeError, match="OPENAI_API_KEY"):
        await svc._call_llm_json([{"role": "user", "content": "x"}])


def test_parse_leading_json_with_trailing_prose_containing_stray_brace():
    # A valid leading object followed by prose that itself contains a '}'.
    # The greedy {.*} regex would swallow the prose and fail; raw_decode
    # parses the leading object and ignores the rest.
    raw = (
        '{"matched": true, "document_id": "D1", "confidence": 0.9, '
        '"matched_signals": [], "reason": "ok"} some prose with a } stray'
    )
    r = parse_classification(raw)
    assert r.document_id == "D1"
    assert r.matched is True
    assert r.error is None


def test_predicted_label_returns_document_id_or_none():
    assert svc.predicted_label(
        ClassificationResult(matched=True, document_id="D1")
    ) == "D1"
    assert svc.predicted_label(
        ClassificationResult(matched=False, document_id=None)
    ) == "NONE"
