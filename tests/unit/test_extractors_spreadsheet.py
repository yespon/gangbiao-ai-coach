# T3-2: spreadsheet merged-cell extraction should preserve repeated values.

from io import BytesIO

from openpyxl import Workbook

from app.extractors.spreadsheet import _extract_xlsx_text


def test_extract_xlsx_fills_merged_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "任务表"
    ws["A1"] = "目的:"
    ws["B1"] = "提升团队协作"
    ws["A2"] = "成果:"
    ws["B2"] = "预算可控"

    ws["A4"] = "核心任务"
    ws.merge_cells("A4:A5")
    ws["B4"] = "拆分目标"
    ws["B5"] = "执行追踪"

    buf = BytesIO()
    wb.save(buf)

    text = _extract_xlsx_text(buf.getvalue())

    assert "[Sheet] 任务表" in text
    assert "任务目的：" in text
    assert "任务成果（预算、交期、完成度）：" in text
    # Merged value from A4 should appear on both logical rows after fill.
    assert "核心任务\t拆分目标" in text
    assert "核心任务\t执行追踪" in text


def test_extract_xlsx_maps_alias_header_to_gangbiao_columns():
    wb = Workbook()
    ws = wb.active
    ws.title = "岗标"
    ws.append(["序号", "岗位任务", "客户价值", "目的", "成果"])
    ws.append([1, "制定标准", "提升客户满意", "统一执行", "通过率提升"])

    buf = BytesIO()
    wb.save(buf)

    text = _extract_xlsx_text(buf.getvalue())

    assert "[Structured] 识别列: 岗位价值->C列" in text
    assert "[Structured] 识别列:" in text
    assert "岗位价值\t岗位任务\t任务目的\t任务成果" in text
    assert "提升客户满意\t制定标准\t统一执行\t通过率提升" in text


def test_extract_xlsx_supports_two_row_headers():
    wb = Workbook()
    ws = wb.active
    ws.title = "双层表头"
    ws.append(["岗位", "岗位", "任务", "任务"])
    ws.append(["价值", "任务", "目的", "成果"])
    ws.append(["支撑组织增长", "梳理流程", "统一口径", "结果可衡量"])

    buf = BytesIO()
    wb.save(buf)

    text = _extract_xlsx_text(buf.getvalue())

    assert "[Structured] 识别列: 岗位价值->A列, 岗位任务->B列, 任务目的->C列, 任务成果->D列" in text
    assert "岗位价值\t岗位任务\t任务目的\t任务成果" in text
    assert "支撑组织增长\t梳理流程\t统一口径\t结果可衡量" in text


def test_extract_xlsx_reports_missing_gangbiao_columns():
    wb = Workbook()
    ws = wb.active
    ws.title = "缺列"
    ws.append(["序号", "事项", "说明"])
    ws.append([1, "A", "B"])

    buf = BytesIO()
    wb.save(buf)

    text = _extract_xlsx_text(buf.getvalue())

    assert "[Structured] 未识别列: 岗位价值" in text


def test_extract_xlsx_does_not_truncate_after_1000_rows():
    wb = Workbook()
    ws = wb.active
    ws.title = "大表"
    ws.append(["岗位任务", "客户价值", "目的", "成果"])
    for idx in range(1, 1206):
        ws.append([f"任务{idx}", f"价值{idx}", f"目的{idx}", f"成果{idx}"])

    buf = BytesIO()
    wb.save(buf)

    text = _extract_xlsx_text(buf.getvalue())

    # Ensure rows after the old 1000-row hard limit are still present.
    assert "任务1205\t价值1205\t目的1205\t成果1205" in text
    assert "[Raw] ..." not in text
