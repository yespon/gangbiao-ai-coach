"""Template classifier service.

Classifies an uploaded Excel file into one of 7 known document templates
(D1..D7) using the ``template_classifier`` prompt. Provides a dedicated
Excel→cells extractor (no gangbiao label rewriting), a one-shot JSON LLM
call at temperature 0, and a tolerant JSON parser.
"""

import json
import re
from io import BytesIO
from typing import Any

import httpx
import xlrd
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from app.core.config import settings
from app.services.prompts import template_classifier as CLASSIFIER_PROMPT

_VALID_DOCUMENT_IDS = {"D1", "D2", "D3", "D4", "D5", "D6", "D7", "D8", "D9"}


class ClassificationResult(BaseModel):
    matched: bool = False
    document_id: str | None = None
    version: str | None = None
    stage: str | None = None
    confidence: float = 0.0
    matched_signals: list[str] = Field(default_factory=list)
    reason: str = ""
    error: str | None = None


def _coerce(obj: dict[str, Any]) -> ClassificationResult:
    """Normalize a parsed JSON object into a validated ClassificationResult."""
    raw_id = obj.get("document_id")
    if isinstance(raw_id, str):
        doc_id = raw_id.strip().upper()
    else:
        doc_id = None
    if doc_id not in _VALID_DOCUMENT_IDS:
        doc_id = None

    # document_id is the source of truth for the label: reconcile `matched`.
    matched = doc_id is not None

    try:
        confidence = float(obj.get("confidence", 0.0))
    except (TypeError, ValueError):
        confidence = 0.0
    confidence = max(0.0, min(1.0, confidence))

    signals = obj.get("matched_signals") or []
    if not isinstance(signals, list):
        signals = []
    signals = [str(s) for s in signals]

    version = obj.get("version")
    stage = obj.get("stage")
    reason = obj.get("reason")

    return ClassificationResult(
        matched=matched,
        document_id=doc_id,
        version=None if version is None else str(version),
        stage=None if stage is None else str(stage),
        confidence=confidence,
        matched_signals=signals,
        reason="" if reason is None else str(reason),
        error=None,
    )


def parse_classification(raw: str | None) -> ClassificationResult:
    """Parse the model's raw text output into a ClassificationResult.

    Tolerant of ```json fences and surrounding prose. Never raises — on
    parse failure returns a result with ``matched=False`` and an ``error``.
    Uses raw_decode so a leading JSON object followed by trailing prose
    (even prose containing a stray ``}``) is still parsed correctly.
    """
    text = (raw or "").strip()

    # Strip a ```json ... ``` fenced block, taking everything between the fences.
    fenced = re.search(r"```(?:json)?\s*(.*?)\s*```", text, re.DOTALL)
    if fenced:
        text = fenced.group(1).strip()

    # Parse the leading JSON object starting at the first '{', ignoring any
    # trailing prose (raw_decode stops at the object's closing brace).
    brace = text.find("{")
    if brace == -1:
        return ClassificationResult(
            matched=False,
            document_id=None,
            reason="解析失败: 未找到 JSON 对象",
            error="no JSON object found",
        )

    try:
        obj, _ = json.JSONDecoder().raw_decode(text[brace:])
    except Exception as exc:  # noqa: BLE001 — any parse failure is a handled error
        return ClassificationResult(
            matched=False,
            document_id=None,
            reason=f"解析失败: {exc}",
            error=str(exc),
        )

    if not isinstance(obj, dict):
        return ClassificationResult(
            matched=False,
            document_id=None,
            reason="解析失败: 模型输出不是 JSON 对象",
            error="non-object JSON",
        )

    return _coerce(obj)


def _normalize_cell(value: Any) -> str:
    """Normalize a raw cell value to the string form used for classification.

    Float integers (e.g. 5.0) become "5"; everything else is str()'d with
    only leading/trailing whitespace trimmed (internal spaces preserved so
    merged-cell fingerprints survive).
    """
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _format_cells(
    rows: int, cols: int, cells: list[tuple[int, int, str]]
) -> str:
    lines = [f"文件尺寸：{rows}行 × {cols}列", "单元格内容（[行,列] 内容）："]
    lines.extend(f"[{r},{c}] {v}" for r, c, v in cells)
    return "\n".join(lines)


def _collect_xlsx(
    raw_bytes: bytes,
) -> tuple[int, int, list[tuple[int, int, str]]]:
    """Return (rows, cols, non-empty [(row, col, value)]) for the first
    non-empty visible worksheet. Hidden sheets are skipped; merged non-anchor
    cells read as None and are skipped, so only the merged top-left value
    appears (no fill). If every visible sheet is empty, fall back to the
    first worksheet so extraction never silently yields nothing.
    """
    wb = load_workbook(filename=BytesIO(raw_bytes), data_only=False, read_only=False)

    sheet = None
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        if any(cell.value not in (None, "") for row in ws.iter_rows() for cell in row):
            sheet = ws
            break
    # No visible sheet has content: fall back to the first non-empty sheet
    # regardless of visibility so extraction never silently yields nothing.
    if sheet is None:
        for ws in wb.worksheets:
            if any(cell.value not in (None, "") for row in ws.iter_rows() for cell in row):
                sheet = ws
                break
    if sheet is None:
        sheet = wb.worksheets[0]

    rows = sheet.max_row or 0
    cols = sheet.max_column or 0
    cells: list[tuple[int, int, str]] = []
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            value = sheet.cell(row=r, column=c).value
            if value in (None, ""):
                continue
            normalized = _normalize_cell(value)
            if normalized:
                cells.append((r, c, normalized))
    return rows, cols, cells


def _collect_xls(
    raw_bytes: bytes,
) -> tuple[int, int, list[tuple[int, int, str]]]:
    """Same shape as _collect_xlsx but for legacy .xls via xlrd.

    xlrd reads merged non-anchor cells as '' (empty), so only the merged
    top-left value appears naturally. Hidden sheets (visibility != 0) are
    skipped; if no visible sheet has content, fall back to the first sheet.
    """
    book = xlrd.open_workbook(file_contents=raw_bytes)
    sheet = None
    for ws in book.sheets():
        if getattr(ws, "visibility", 0) != 0:
            continue
        if any(ws.cell_value(r, c) not in (None, "") for r in range(ws.nrows) for c in range(ws.ncols)):
            sheet = ws
            break
    # No visible sheet has content: fall back to the first non-empty sheet
    # regardless of visibility so extraction never silently yields nothing.
    if sheet is None:
        for ws in book.sheets():
            if any(ws.cell_value(r, c) not in (None, "") for r in range(ws.nrows) for c in range(ws.ncols)):
                sheet = ws
                break
    if sheet is None:
        sheet = book.sheets()[0]
    rows = sheet.nrows
    cols = sheet.ncols
    cells: list[tuple[int, int, str]] = []
    for r in range(rows):
        for c in range(cols):
            value = sheet.cell_value(r, c)
            if value in (None, ""):
                continue
            normalized = _normalize_cell(value)
            if normalized:
                cells.append((r + 1, c + 1, normalized))  # 1-based for output
    return rows, cols, cells


def extract_cells_for_classification(raw_bytes: bytes, ext: str) -> str:
    """Extract an Excel file into the prompt's expected input format.

    ``ext`` is the lowercased file extension including the dot, e.g. ".xlsx".
    Raises RuntimeError for unsupported extensions.
    """
    ext = (ext or "").lower()
    if ext == ".xlsx":
        rows, cols, cells = _collect_xlsx(raw_bytes)
    elif ext == ".xls":
        rows, cols, cells = _collect_xls(raw_bytes)
    else:
        raise RuntimeError(f"不支持的文件扩展名: {ext!r}（仅支持 .xlsx / .xls）")
    return _format_cells(rows, cols, cells)


async def _call_llm_json(messages: list[dict[str, str]]) -> str:
    """One-shot LLM call returning the raw message content (temperature 0).

    Raises RuntimeError if the API key is missing or the response is an error.
    """
    api_key = settings.openai_api_key.strip()
    if not api_key:
        raise RuntimeError("未配置 OPENAI_API_KEY，无法调用模板分类器")

    model = settings.openai_model
    base_url = settings.openai_base_url.rstrip("/")
    url = f"{base_url}/chat/completions"
    payload = {"model": model, "messages": messages, "temperature": 0}
    headers = {"Authorization": f"Bearer {api_key}"}

    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.post(url, json=payload, headers=headers)
        if response.status_code >= 400:
            raise RuntimeError(
                f"LLM 调用失败: {response.status_code} {response.text}"
            )
        data = response.json()

    choices = data.get("choices") or []
    if not choices:
        raise RuntimeError(f"LLM 返回无 choices: {data}")
    content = choices[0].get("message", {}).get("content")
    return content or ""


async def classify_text(text: str) -> ClassificationResult:
    """Classify pre-extracted cell text into a template."""
    messages = [
        {"role": "system", "content": CLASSIFIER_PROMPT},
        {"role": "user", "content": text},
    ]
    raw = await _call_llm_json(messages)
    return parse_classification(raw)


async def classify_file(raw_bytes: bytes, ext: str) -> ClassificationResult:
    """Extract an Excel file and classify it."""
    text = extract_cells_for_classification(raw_bytes, ext)
    return await classify_text(text)


def predicted_label(result: ClassificationResult) -> str:
    """Map a result to an eval label: the document_id, or 'NONE' when unmatched."""
    return result.document_id if result.document_id else "NONE"
