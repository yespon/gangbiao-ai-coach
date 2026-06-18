from dataclasses import dataclass
from io import BytesIO
from typing import Any, Literal
import uuid

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy.sql import Select

from app.models.db_models import ManagedUserDB
from app.services.whitelist_service import MAX_WHITELIST_ROWS, WHITELIST_DENY_MESSAGE

ROLE_LABELS = {"管理员": "admin", "教练": "coach", "学员": "student", "admin": "admin", "coach": "coach", "student": "student"}
ENABLED_LABELS = {"启用": True, "禁用": False, "true": True, "false": False, "是": True, "否": False}
MANAGED_USER_TEMPLATE_HEADERS = ["工号", "姓名", "邮箱", "一级部门", "主角色", "兼任教练", "所属教练工号", "启用状态"]


@dataclass
class ManagedUserParseResult:
    rows: list[dict[str, Any]]
    errors: list[dict[str, Any]]


def _cell_text(value) -> str:
    return str(value).strip() if value is not None else ""


def normalize_employee_no(value: str) -> str:
    return value.strip()


def is_effective_coach(primary_role: str, is_coach: bool) -> bool:
    return primary_role == "coach" or (primary_role == "admin" and is_coach)


def normalize_managed_user_role(
    primary_role: str | None, is_coach: bool, coach_id: str | uuid.UUID | None
):
    role = ROLE_LABELS.get((primary_role or "学员").strip())
    if role is None:
        raise ValueError("主角色必须是管理员、教练或学员")
    if role == "coach":
        return {"primary_role": "coach", "is_coach": True, "coach_id": None}
    if role == "admin":
        return {"primary_role": "admin", "is_coach": bool(is_coach), "coach_id": None}
    return {"primary_role": "student", "is_coach": False, "coach_id": coach_id}


def protect_system_admin_patch(employee_no: str, admin_employee_nos: set[str], requested: dict[str, Any]) -> dict[str, Any]:
    if employee_no not in admin_employee_nos:
        return requested
    protected = dict(requested)
    protected["enabled"] = True
    protected["primary_role"] = "admin"
    return protected


def build_managed_user_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "用户管理"
    ws.append(MANAGED_USER_TEMPLATE_HEADERS)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_indexes(header: list[str]) -> dict[str, int] | None:
    try:
        return {name: header.index(name) for name in MANAGED_USER_TEMPLATE_HEADERS}
    except ValueError:
        return None


def _parse_bool(value: str, default: bool) -> bool | None:
    if not value:
        return default
    lowered = value.lower()
    return ENABLED_LABELS.get(lowered) if lowered in ENABLED_LABELS else ENABLED_LABELS.get(value)


def parse_managed_user_excel(raw: bytes) -> ManagedUserParseResult:
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    try:
        ws = wb.active
        header = [_cell_text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        indexes = _header_indexes(header)
        if indexes is None:
            return ManagedUserParseResult(rows=[], errors=[{"row": 1, "reason": "表头必须包含工号、姓名、邮箱、一级部门、主角色、兼任教练、所属教练工号、启用状态"}])

        latest: dict[str, dict[str, Any]] = {}
        errors: list[dict[str, Any]] = []
        for row_no, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_no - 1 > MAX_WHITELIST_ROWS:
                errors.append({"row": row_no, "reason": f"超过最大导入行数 {MAX_WHITELIST_ROWS}"})
                break
            employee_no = normalize_employee_no(_cell_text(row[indexes["工号"]].value if indexes["工号"] < len(row) else None))
            if not employee_no:
                errors.append({"row": row_no, "reason": "工号为空"})
                continue
            role_text = _cell_text(row[indexes["主角色"]].value if indexes["主角色"] < len(row) else None) or "学员"
            coach_flag_text = _cell_text(row[indexes["兼任教练"]].value if indexes["兼任教练"] < len(row) else None)
            coach_flag = _parse_bool(coach_flag_text, False)
            if coach_flag is None:
                errors.append({"row": row_no, "reason": "兼任教练必须是是或否"})
                continue
            try:
                normalized = normalize_managed_user_role(role_text, coach_flag, None)
            except ValueError as exc:
                errors.append({"row": row_no, "reason": str(exc)})
                continue
            if normalized["primary_role"] == "student" and coach_flag:
                errors.append({"row": row_no, "reason": "学员不能兼任教练"})
                continue
            enabled_text = _cell_text(row[indexes["启用状态"]].value if indexes["启用状态"] < len(row) else None)
            enabled = _parse_bool(enabled_text, True)
            if enabled is None:
                errors.append({"row": row_no, "reason": "启用状态必须是启用或禁用"})
                continue
            if employee_no in latest:
                errors.append({"row": row_no, "reason": "工号重复，已使用最后一条记录"})
            latest[employee_no] = {
                "employee_no": employee_no,
                "name": _cell_text(row[indexes["姓名"]].value if indexes["姓名"] < len(row) else None) or None,
                "email": _cell_text(row[indexes["邮箱"]].value if indexes["邮箱"] < len(row) else None) or None,
                "department_level1": _cell_text(row[indexes["一级部门"]].value if indexes["一级部门"] < len(row) else None) or None,
                "primary_role": normalized["primary_role"],
                "is_coach": normalized["is_coach"],
                "coach_employee_no": _cell_text(row[indexes["所属教练工号"]].value if indexes["所属教练工号"] < len(row) else None) or None,
                "enabled": enabled,
                "row": row_no,
            }
        return ManagedUserParseResult(rows=list(latest.values()), errors=errors)
    finally:
        wb.close()


def resolve_import_coach_links(rows: list[dict[str, Any]], existing_coach_employee_nos: set[str]) -> list[dict[str, Any]]:
    batch_coaches = {row["employee_no"] for row in rows if is_effective_coach(row["primary_role"], row["is_coach"])}
    valid_coaches = existing_coach_employee_nos | batch_coaches
    errors: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        coach_employee_no = row.get("coach_employee_no")
        if row["primary_role"] == "student" and coach_employee_no and coach_employee_no not in valid_coaches:
            errors.append({"row": row.get("row", idx), "reason": "所属教练工号不存在或不是教练"})
    return errors


async def get_managed_user_by_employee_no(db: AsyncSession, employee_no: str) -> ManagedUserDB | None:
    result = await db.execute(select(ManagedUserDB).where(ManagedUserDB.employee_no == employee_no))
    return result.scalar_one_or_none()


async def ensure_managed_user_allowed(db: AsyncSession, employee_no: str, is_system_admin: bool) -> ManagedUserDB:
    profile = await get_managed_user_by_employee_no(db, employee_no)
    if profile is None and is_system_admin:
        profile = ManagedUserDB(
            employee_no=employee_no,
            primary_role="admin",
            is_coach=False,
            enabled=True,
            source="system",
        )
        db.add(profile)
        await db.commit()
        await db.refresh(profile)
        return profile
    if profile is None:
        raise PermissionError(WHITELIST_DENY_MESSAGE)
    if not profile.enabled and not is_system_admin:
        raise PermissionError(WHITELIST_DENY_MESSAGE)
    if is_system_admin and (profile.primary_role != "admin" or not profile.enabled):
        profile.primary_role = "admin"
        profile.enabled = True
        await db.commit()
        await db.refresh(profile)
    return profile


async def existing_coach_employee_nos(db: AsyncSession) -> set[str]:
    result = await db.execute(select(ManagedUserDB).where(ManagedUserDB.primary_role.in_(["admin", "coach"])))
    profiles = result.scalars().all()
    return {p.employee_no for p in profiles if is_effective_coach(p.primary_role, p.is_coach)}


async def upsert_managed_user(
    db: AsyncSession,
    payload: dict[str, Any],
    source: str,
    created_by: uuid.UUID | None,
    admin_employee_nos: set[str],
) -> tuple[ManagedUserDB, bool]:
    employee_no = normalize_employee_no(payload["employee_no"])
    requested = protect_system_admin_patch(employee_no, admin_employee_nos, dict(payload))
    normalized = normalize_managed_user_role(requested.get("primary_role"), requested.get("is_coach", False), requested.get("coach_id"))
    requested.update(normalized)
    result = await db.execute(select(ManagedUserDB).where(ManagedUserDB.employee_no == employee_no))
    profile = result.scalar_one_or_none()
    created = profile is None
    if profile is None:
        profile = ManagedUserDB(employee_no=employee_no, created_by=created_by)
        db.add(profile)
    profile.name = requested.get("name")
    profile.email = requested.get("email")
    profile.department_level1 = requested.get("department_level1")
    profile.primary_role = requested["primary_role"]
    profile.is_coach = requested["is_coach"]
    profile.coach_id = requested.get("coach_id") if requested["primary_role"] == "student" else None
    profile.enabled = requested.get("enabled", True)
    profile.source = source
    return profile, created


ManagedUserCoachFilter = Literal["all", "unassigned"] | str  # "all" | "unassigned" | "<uuid>"


@dataclass(slots=True)
class ManagedUserListFilters:
    q: str | None = None
    role: str | None = None
    enabled: bool | None = None
    coach_filter: str = "all"  # "all" | "unassigned" | "<uuid>"
    department_level1: str | None = None
    has_email: bool | None = None


def _haystack_columns() -> list:
    return [
        ManagedUserDB.employee_no,
        ManagedUserDB.name,
        ManagedUserDB.email,
        ManagedUserDB.department_level1,
    ]


def build_managed_user_filtered_stmt(filters: ManagedUserListFilters) -> Select:
    """Build a SELECT that already has WHERE applied; callers add LIMIT/OFFSET/ORDER BY."""
    stmt = select(ManagedUserDB).options(selectinload(ManagedUserDB.coach))
    if filters.role:
        stmt = stmt.where(ManagedUserDB.primary_role == filters.role)
    if filters.enabled is not None:
        stmt = stmt.where(ManagedUserDB.enabled.is_(filters.enabled))
    if filters.department_level1:
        stmt = stmt.where(ManagedUserDB.department_level1 == filters.department_level1)
    if filters.has_email is True:
        stmt = stmt.where(ManagedUserDB.email.is_not(None))
    elif filters.has_email is False:
        stmt = stmt.where(ManagedUserDB.email.is_(None))
    if filters.coach_filter == "unassigned":
        stmt = stmt.where(ManagedUserDB.coach_id.is_(None))
    elif filters.coach_filter not in (None, "all", ""):
        try:
            stmt = stmt.where(ManagedUserDB.coach_id == uuid.UUID(filters.coach_filter))
        except ValueError:
            # Invalid UUIDs are treated as "no match" — caller can decide to 400.
            stmt = stmt.where(False)
    if filters.q:
        needle = f"%{filters.q.strip().lower()}%"
        cols = _haystack_columns()
        stmt = stmt.where(or_(*[func.lower(c).like(needle) for c in cols]))
    return stmt
