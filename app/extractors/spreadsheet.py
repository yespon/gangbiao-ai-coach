import re
import os
from io import BytesIO
from typing import Any

import xlrd
from openpyxl import load_workbook


_HEADER_SCAN_LIMIT = 12
_PREVIEW_ROW_LIMIT = 12
_GANGBIAO_COLUMN_ORDER = ("岗位价值", "岗位任务", "任务目的", "任务成果")
_GANGBIAO_HEADER_ALIASES = {
    "岗位价值": {
        "岗位价值",
        "价值",
        "岗位核心价值",
        "岗位价值点",
        "岗位价值输出",
        "岗位价值说明",
        "岗位价值描述",
        "客户价值",
        "价值贡献",
        "岗值",
    },
    "岗位任务": {
        "岗位任务",
        "任务",
        "核心任务",
        "关键任务",
        "主要任务",
        "工作任务",
        "任务事项",
        "岗位职责",
        "职责",
    },
    "任务目的": {
        "任务目的",
        "目的",
        "任务目标",
        "目标",
    },
    "任务成果": {
        "任务成果",
        "成果",
        "交付成果",
        "输出成果",
        "产出",
    },
}


def _int_env(name: str, default: int, minimum: int) -> int:
    raw = os.getenv(name, "").strip()
    if not raw:
        return default
    try:
        value = int(raw)
    except ValueError:
        return default
    return max(value, minimum)


# 0 means no hard limit so the full sheet can be extracted.
_RAW_ROW_LIMIT = _int_env("SPREADSHEET_RAW_ROW_LIMIT", 0, 0)
_RAW_COL_LIMIT = _int_env("SPREADSHEET_RAW_COL_LIMIT", 0, 0)


def _effective_limit(available: int, configured_limit: int) -> int:
    if configured_limit <= 0:
        return max(available, 1)
    return min(max(available, 1), configured_limit)


def _normalize_cell_value(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_gangbiao_labels(text: str) -> str:
    if not text:
        return text

    normalized_lines: list[str] = []
    for raw_line in text.splitlines():
        line = raw_line

        # Normalize common purpose label variants.
        line = re.sub(r"\b目的\s*[:：]", "任务目的：", line)

        # Normalize result label variants while keeping the remaining sentence.
        line = re.sub(
            r"\b成果\s*(?:（[^）]*）)?\s*[:：]",
            "任务成果（预算、交期、完成度）：",
            line,
        )

        # Keep standardized labels deterministic by trimming spaces after colon.
        line = re.sub(r"任务目的：\s*", "任务目的：", line)
        line = re.sub(r"任务成果（预算、交期、完成度）：\s*", "任务成果（预算、交期、完成度）：", line)

        normalized_lines.append(line)

    return "\n".join(normalized_lines)


def _compact_label(text: str) -> str:
    if not text:
        return ""
    compact = str(text).strip().lower()
    compact = re.sub(r"[\s\t\r\n\u3000]+", "", compact)
    compact = re.sub(r"[：:，,。\.；;、\-_/\\|（）()【】\[\]《》<>]", "", compact)
    return compact


def _excel_col_name(index: int) -> str:
    # 0-based column index to Excel name (A, B, ..., AA, AB, ...)
    n = index + 1
    parts: list[str] = []
    while n > 0:
        n -= 1
        parts.append(chr(ord("A") + (n % 26)))
        n //= 26
    return "".join(reversed(parts))


def _map_header_to_canonical(label: str) -> str | None:
    normalized = _compact_label(label)
    if not normalized:
        return None

    for canonical, aliases in _GANGBIAO_HEADER_ALIASES.items():
        for alias in aliases:
            alias_norm = _compact_label(alias)
            if not alias_norm:
                continue
            if alias_norm in normalized:
                return canonical
    return None


def _find_header_mapping(rows: list[list[str]]) -> tuple[int | None, dict[str, int], bool]:
    scan_rows = min(len(rows), _HEADER_SCAN_LIMIT)
    best_score = -1
    best_header_idx: int | None = None
    best_mapping: dict[str, int] = {}
    best_uses_next_row = False

    for row_idx in range(scan_rows):
        row = rows[row_idx]
        next_row = rows[row_idx + 1] if row_idx + 1 < len(rows) else []
        width = max(len(row), len(next_row))

        mapping_single: dict[str, int] = {}
        mapping_combined: dict[str, int] = {}

        for col_idx in range(width):
            cell = row[col_idx] if col_idx < len(row) else ""
            mapped_single = _map_header_to_canonical(cell)
            if mapped_single and mapped_single not in mapping_single:
                mapping_single[mapped_single] = col_idx

            if next_row:
                cell_next = next_row[col_idx] if col_idx < len(next_row) else ""
                merged_cell = f"{cell}{cell_next}"
                mapped_combined = _map_header_to_canonical(merged_cell)
                if mapped_combined and mapped_combined not in mapping_combined:
                    mapping_combined[mapped_combined] = col_idx

        score_single = len(mapping_single)
        score_combined = len(mapping_combined)

        if score_combined > score_single:
            score = score_combined
            mapping = mapping_combined
            uses_next_row = True
        else:
            score = score_single
            mapping = mapping_single
            uses_next_row = False

        if score > best_score:
            best_score = score
            best_header_idx = row_idx
            best_mapping = mapping
            best_uses_next_row = uses_next_row

    if best_score <= 0:
        return None, {}, False

    return best_header_idx, best_mapping, best_uses_next_row


def _build_structured_sheet_preview(sheet_name: str, rows: list[list[str]]) -> list[str]:
    lines: list[str] = [f"[Sheet] {sheet_name}"]
    if not rows:
        lines.append("[Structured] 空表")
        return lines

    header_idx, mapping, uses_next_row = _find_header_mapping(rows)
    if header_idx is None:
        lines.append("[Structured] 未识别到岗位标准化表头")
        lines.append("[Structured] 未识别列: " + ", ".join(_GANGBIAO_COLUMN_ORDER))
        return lines

    lines.append(f"[Structured] 识别表头行: 第{header_idx + 1}行")
    if uses_next_row:
        lines.append(f"[Structured] 表头跨行合并: 第{header_idx + 1}-{header_idx + 2}行")

    recognized = [
        f"{name}->{_excel_col_name(mapping[name])}列"
        for name in _GANGBIAO_COLUMN_ORDER
        if name in mapping
    ]
    missing = [name for name in _GANGBIAO_COLUMN_ORDER if name not in mapping]

    if recognized:
        lines.append("[Structured] 识别列: " + ", ".join(recognized))
    if missing:
        lines.append("[Structured] 未识别列: " + ", ".join(missing))

    data_start = header_idx + (2 if uses_next_row else 1)
    preview_rows: list[list[str]] = []
    for row in rows[data_start:]:
        values = [row[mapping[name]] if name in mapping and mapping[name] < len(row) else "" for name in _GANGBIAO_COLUMN_ORDER]
        if any(value.strip() for value in values):
            preview_rows.append(values)
        if len(preview_rows) >= _PREVIEW_ROW_LIMIT:
            break

    lines.append("[Structured] 关键列预览:")
    lines.append("\t".join(_GANGBIAO_COLUMN_ORDER))
    if preview_rows:
        for values in preview_rows:
            lines.append("\t".join(values))
    else:
        lines.append("(无数据行)")

    return lines


def _rows_to_raw_lines(rows: list[list[str]]) -> list[str]:
    lines: list[str] = []
    for row in rows:
        if not any(cell for cell in row):
            continue
        trimmed = list(row)
        while trimmed and trimmed[-1] == "":
            trimmed.pop()
        lines.append("\t".join(trimmed))
    return lines


def _extract_xlsx_text(raw_bytes: bytes) -> str:
    try:
        # Some files have incorrect dimension metadata (e.g. A1), which breaks
        # read_only iteration. Normal mode is more robust for these workbooks.
        wb = load_workbook(filename=BytesIO(raw_bytes), data_only=False, read_only=False)
    except Exception:
        return ""

    lines: list[str] = []

    for sheet in wb.worksheets:
        rows: list[list[str]] = []
        merged_values: dict[tuple[int, int], Any] = {}
        for merged in sheet.merged_cells.ranges:
            top_left = sheet.cell(row=merged.min_row, column=merged.min_col).value
            if top_left in (None, ""):
                continue
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    merged_values[(r, c)] = top_left

        max_rows = _effective_limit(sheet.max_row, _RAW_ROW_LIMIT)
        max_cols = _effective_limit(sheet.max_column, _RAW_COL_LIMIT)

        for row_idx in range(1, max_rows + 1):
            cells: list[str] = []
            row_has_value = False

            for col_idx in range(1, max_cols + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if value in (None, ""):
                    value = merged_values.get((row_idx, col_idx))

                normalized = _normalize_cell_value(value) if value not in (None, "") else ""
                if normalized:
                    row_has_value = True
                cells.append(normalized)

            if row_has_value:
                while cells and cells[-1] == "":
                    cells.pop()
                rows.append(cells)

        lines.extend(_build_structured_sheet_preview(sheet.title, rows))
        lines.append("[Raw]")
        lines.extend(_rows_to_raw_lines(rows))

        if sheet.max_row > max_rows or sheet.max_column > max_cols:
            lines.append("[Raw] ...")

    return _normalize_gangbiao_labels("\n".join(lines))


def _extract_xls_text(raw_bytes: bytes) -> str:
    try:
        book = xlrd.open_workbook(file_contents=raw_bytes)
    except Exception:
        return ""

    lines: list[str] = []

    for sheet in book.sheets():
        rows: list[list[str]] = []
        max_rows = _effective_limit(sheet.nrows, _RAW_ROW_LIMIT)
        max_cols = _effective_limit(sheet.ncols, _RAW_COL_LIMIT)

        merged_values: dict[tuple[int, int], Any] = {}
        for rlo, rhi, clo, chi in getattr(sheet, "merged_cells", []):
            top_left = sheet.cell_value(rlo, clo)
            if top_left in (None, ""):
                continue
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    merged_values[(r, c)] = top_left

        for r in range(max_rows):
            cells: list[str] = []
            row_has_value = False
            for c in range(max_cols):
                value = sheet.cell_value(r, c)
                if value in (None, ""):
                    value = merged_values.get((r, c))

                normalized = _normalize_cell_value(value) if value not in (None, "") else ""
                if normalized:
                    row_has_value = True
                cells.append(normalized)

            if row_has_value:
                while cells and cells[-1] == "":
                    cells.pop()
                rows.append(cells)

        lines.extend(_build_structured_sheet_preview(sheet.name, rows))
        lines.append("[Raw]")
        lines.extend(_rows_to_raw_lines(rows))

        if sheet.nrows > max_rows or sheet.ncols > max_cols:
            lines.append("[Raw] ...")

    return _normalize_gangbiao_labels("\n".join(lines))
